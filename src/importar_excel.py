"""
ETAPA 1B — IMPORTAÇÃO DA PLANILHA EXCEL
========================================

Por que este módulo existe:
Baixar 3.700 concursos da API da Caixa leva ~30 minutos. Se você já tem o
histórico em Excel, importamos em 2 segundos e usamos a API apenas para os
concursos novos, do dia a dia em diante.

O que ele faz:
1. Lê a planilha e DETECTA SOZINHO o formato das dezenas (3 formatos comuns).
2. Converte tudo para o mesmo JSON padrão usado pelo resto do projeto.
3. AUDITA a base: concursos faltando, duplicados, linhas com menos de 15
   dezenas, dezenas fora de 1-25, datas inválidas.
4. Grava em data/lotofacil.json.

Como rodar:
    python -m src.importar_excel "caminho/para/Lotofacil.xlsx"
    python -m src.importar_excel "planilha.xlsx" --aba "LOTOFÁCIL"
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src import coleta  # reutiliza salvar_base / validar_concurso  # noqa: E402


# ---------------------------------------------------------------------------
# AJUDANTES DE LEITURA
# ---------------------------------------------------------------------------


def _texto(valor: Any) -> str:
    return "" if valor is None else str(valor).strip()


def _normalizar_cabecalho(valor: Any) -> str:
    """Deixa o cabeçalho comparável: minúsculo, sem acento, sem espaço extra."""
    t = _texto(valor).lower()
    trocas = {"á": "a", "â": "a", "ã": "a", "à": "a", "é": "e", "ê": "e",
              "í": "i", "ó": "o", "ô": "o", "õ": "o", "ú": "u", "ç": "c"}
    for de, para in trocas.items():
        t = t.replace(de, para)
    return re.sub(r"\s+", " ", t).strip()


def _para_int(valor: Any) -> int | None:
    """Converte '07', 7, 7.0 para 7. Devolve None se não for número."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return int(valor)
    t = _texto(valor)
    return int(t) if t.isdigit() else None


def _para_dinheiro(valor: Any) -> float:
    """Converte 'R$49.765,82' (ou 49765.82) para 49765.82."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    t = _texto(valor).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def _para_data(valor: Any) -> str:
    """Converte data do Excel (texto ou datetime) para 'AAAA-MM-DD'."""
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%Y-%m-%d")
    t = _texto(valor)
    for formato in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(t, formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return t  # devolve como veio; a auditoria vai sinalizar


# ---------------------------------------------------------------------------
# DETECÇÃO DO FORMATO
# ---------------------------------------------------------------------------

# Formato A: uma coluna por bola   -> Bola1 ... Bola15
# Formato B: dezenas numa célula   -> "01-03-07-09..."
# Formato C: matriz marcada        -> colunas 1..25 com X / 1 / a própria dezena


def detectar_layout(cabecalho: list[Any]) -> dict[str, Any]:
    limpo = [_normalizar_cabecalho(c) for c in cabecalho]

    col_concurso = next(
        (i for i, c in enumerate(limpo) if c.startswith("concurso") or c in ("n concurso", "sorteio")),
        None,
    )
    col_data = next((i for i, c in enumerate(limpo) if c.startswith("data")), None)
    # Coluna opcional: quantos acertaram 15 dezenas (vira prova social no conteúdo)
    col_ganhadores = next(
        (i for i, c in enumerate(limpo) if c.startswith("ganhadores 15")), None
    )
    # Colunas opcionais de rateio por faixa — permitem calcular retorno exato
    # (os valores dos prêmios mudaram ao longo dos anos)
    col_rateios = {
        faixa: next((i for i, c in enumerate(limpo) if c.startswith(f"rateio {faixa}")), None)
        for faixa in (11, 12, 13, 14, 15)
    }

    # Formato A — procura colunas tipo "bola1", "bola 1", "b1", "dezena 1", "n1"
    colunas_bola = [
        i for i, c in enumerate(limpo)
        if re.fullmatch(r"(bola|b|dezena|d|n|num|numero)\s*0?([1-9]|1[0-5])", c)
    ]
    if len(colunas_bola) >= 15:
        return {
            "tipo": "colunas_por_bola",
            "col_concurso": col_concurso,
            "col_data": col_data,
            "col_ganhadores": col_ganhadores,
            "col_rateios": col_rateios,
            "colunas": colunas_bola[:15],
        }

    # Formato C — colunas chamadas 1, 2, 3, ..., 25
    colunas_matriz = {
        int(c): i for i, c in enumerate(limpo)
        if c.isdigit() and 1 <= int(c) <= 25
    }
    if len(colunas_matriz) >= 25:
        return {
            "tipo": "matriz",
            "col_concurso": col_concurso,
            "col_data": col_data,
            "col_ganhadores": col_ganhadores,
            "col_rateios": col_rateios,
            "colunas": colunas_matriz,
        }

    # Formato B — alguma coluna com "dezena(s)", "numeros", "resultado"
    col_texto = next(
        (i for i, c in enumerate(limpo)
         if any(p in c for p in ("dezenas", "numeros", "resultado", "sorteadas"))),
        None,
    )
    if col_texto is not None:
        return {
            "tipo": "celula_unica",
            "col_concurso": col_concurso,
            "col_data": col_data,
            "col_ganhadores": col_ganhadores,
            "col_rateios": col_rateios,
            "coluna": col_texto,
        }

    raise ValueError(
        "Não reconheci o formato da planilha. Cabeçalho encontrado: " + ", ".join(limpo[:20])
    )


def extrair_dezenas(linha: tuple, layout: dict[str, Any]) -> list[int]:
    tipo = layout["tipo"]

    if tipo == "colunas_por_bola":
        valores = [_para_int(linha[i]) if i < len(linha) else None for i in layout["colunas"]]
        return sorted(v for v in valores if v is not None)

    if tipo == "matriz":
        marcadas = []
        for dezena, idx in layout["colunas"].items():
            if idx >= len(linha):
                continue
            celula = _texto(linha[idx]).lower()
            if celula not in ("", "0", "none", "-"):
                marcadas.append(dezena)
        return sorted(marcadas)

    # celula_unica
    idx = layout["coluna"]
    bruto = _texto(linha[idx]) if idx < len(linha) else ""
    return sorted(int(n) for n in re.findall(r"\d{1,2}", bruto))


# ---------------------------------------------------------------------------
# IMPORTAÇÃO + AUDITORIA
# ---------------------------------------------------------------------------


def importar(caminho: str | Path, aba: str | None = None, verboso: bool = True) -> dict[str, Any]:
    caminho = Path(caminho)
    if not caminho.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {caminho}")

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]

    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas)
    layout = detectar_layout(list(cabecalho))

    if verboso:
        print(f"Aba lida: {ws.title}")
        print(f"Formato detectado: {layout['tipo']}")

    registros: list[dict[str, Any]] = []
    problemas: dict[str, list] = {
        "sem_15_dezenas": [], "dezena_fora_faixa": [],
        "dezena_repetida": [], "data_invalida": [], "sem_concurso": [],
    }

    for numero_linha, linha in enumerate(linhas, start=2):
        if linha is None or all(c is None or _texto(c) == "" for c in linha):
            continue  # linha em branco

        concurso = _para_int(linha[layout["col_concurso"]]) if layout["col_concurso"] is not None else None
        if concurso is None:
            problemas["sem_concurso"].append(numero_linha)
            continue

        dezenas = extrair_dezenas(linha, layout)

        if len(dezenas) != 15:
            problemas["sem_15_dezenas"].append(concurso)
            continue
        if len(set(dezenas)) != 15:
            problemas["dezena_repetida"].append(concurso)
            continue
        if not all(1 <= d <= 25 for d in dezenas):
            problemas["dezena_fora_faixa"].append(concurso)
            continue

        idx_g = layout.get("col_ganhadores")
        ganhadores = (_para_int(linha[idx_g]) or 0) if idx_g is not None and idx_g < len(linha) else 0

        data = _para_data(linha[layout["col_data"]]) if layout["col_data"] is not None else ""
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", data):
            problemas["data_invalida"].append(concurso)

        rateios = {
            str(faixa): _para_dinheiro(linha[idx]) if idx is not None and idx < len(linha) else 0.0
            for faixa, idx in (layout.get("col_rateios") or {}).items()
        }

        registros.append({
            "concurso": concurso,
            "data": data,
            "dezenas": dezenas,
            "acumulado": False,
            "ganhadores_15": ganhadores,
            "rateios": rateios,
            "fonte": "excel",
        })

    # ---- Auditoria de integridade da sequência ----
    numeros = sorted(r["concurso"] for r in registros)
    duplicados = sorted({n for n in numeros if numeros.count(n) > 1}) if len(numeros) != len(set(numeros)) else []
    buracos = sorted(set(range(min(numeros), max(numeros) + 1)) - set(numeros)) if numeros else []

    coleta.salvar_base(registros)

    relatorio = {
        "importados": len(registros),
        "primeiro": min(numeros) if numeros else None,
        "ultimo": max(numeros) if numeros else None,
        "buracos": buracos,
        "duplicados": duplicados,
        "problemas": {k: v for k, v in problemas.items() if v},
        "arquivo": str(coleta.ARQUIVO_BASE),
    }

    if verboso:
        print("\n=== AUDITORIA DA BASE ===")
        print(f"Concursos importados : {relatorio['importados']}")
        print(f"Intervalo            : {relatorio['primeiro']} a {relatorio['ultimo']}")
        print(f"Faltando no meio     : {len(buracos)}" + (f" -> {buracos[:15]}" if buracos else " (nenhum)"))
        print(f"Duplicados           : {len(duplicados)}" + (f" -> {duplicados[:15]}" if duplicados else " (nenhum)"))
        for chave, lista in relatorio["problemas"].items():
            print(f"{chave:21}: {len(lista)} -> {lista[:10]}")
        print(f"\nGravado em: {relatorio['arquivo']}")

    return relatorio


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Importa histórico da Lotofácil de uma planilha Excel")
    parser.add_argument("planilha", help="caminho do arquivo .xlsx")
    parser.add_argument("--aba", default=None, help="nome da aba (padrão: a primeira)")
    args = parser.parse_args()
    importar(args.planilha, args.aba)
